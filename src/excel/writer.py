from datetime import date, datetime
from typing import Optional
import openpyxl
from openpyxl.cell.cell import Cell
from ..models.report import DailyEntry
from .finder import CellFinder, find_column_by_header


class ExcelWriteError(Exception):
    pass


class SafeExcelWriter:
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.wb: Optional[openpyxl.Workbook] = None
        self.modified = False

    def open(self):
        self.wb = openpyxl.load_workbook(self.workbook_path)
        return self

    def close(self):
        if self.wb and self.modified:
            self.wb.save(self.workbook_path)
            self.modified = True

    def __enter__(self):
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def safe_write_cell(self, worksheet, row: int, col: int, value) -> bool:
        if row < 1 or col < 1:
            return False

        cell = worksheet.cell(row, col)

        if cell.data_type == 'f':
            return False

        cell.value = value
        self.modified = True
        return True

    def write_daily_entry(
        self,
        worksheet,
        entry: DailyEntry,
        date_col: int = 3,
        km_col: int = 4,
        speed_col: int = 5,
        parking_col: int = 6,
        fuel_col: int = 7
    ) -> bool:
        finder = CellFinder(worksheet)

        row = finder.find_date_row(entry.date, date_col)

        if not row:
            raise ExcelWriteError(f"No se encontró fila para fecha {entry.date}")

        self.safe_write_cell(worksheet, row, km_col, entry.kilometers)
        self.safe_write_cell(worksheet, row, speed_col, entry.speed_excess_minutes)
        self.safe_write_cell(worksheet, row, parking_col, entry.parking_minutes)
        self.safe_write_cell(worksheet, row, fuel_col, entry.fuel_liters)

        return True

    def check_duplicate(
        self,
        worksheet,
        entry_date: date,
        kilometers: float,
        tolerance: float = 0.1
    ) -> bool:
        finder = CellFinder(worksheet)
        row = finder.find_date_row(entry_date)

        if not row:
            return False

        km_cell = worksheet.cell(row, 2).value
        if km_cell is not None:
            try:
                existing_km = float(km_cell)
                if abs(existing_km - kilometers) < tolerance:
                    return True
            except (ValueError, TypeError):
                pass

        return False

    def write_entry_safe(
        self,
        worksheet,
        entry: DailyEntry,
        column_mapping: dict[str, int]
    ) -> bool:
        date_col = column_mapping.get("FECHA", 3)
        km_col = column_mapping.get("KILOMETRAJE", 4)
        speed_col = column_mapping.get("EXCESO VELOCIDAD", 5)
        parking_col = column_mapping.get("ESTACIONAMIENTO", 6)
        fuel_col = column_mapping.get("COMBUSTIBLE", 7)

        if not km_col:
            km_col = 4
        if not speed_col:
            speed_col = 5
        if not parking_col:
            parking_col = 6
        if not fuel_col:
            fuel_col = 7

        finder = CellFinder(worksheet)
        row = finder.find_date_row(entry.date, date_col)

        if not row:
            raise ExcelWriteError(f"No se encontró fila para fecha {entry.date}")

        if self.check_duplicate(worksheet, entry.date, entry.kilometers):
            return False

        self.safe_write_cell(worksheet, row, km_col, entry.kilometers)
        self.safe_write_cell(worksheet, row, speed_col, entry.speed_excess_minutes)
        self.safe_write_cell(worksheet, row, parking_col, entry.parking_minutes)
        self.safe_write_cell(worksheet, row, fuel_col, entry.fuel_liters)

        return True


class ExcelUpdater:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb: Optional[openpyxl.Workbook] = None

    def load(self) -> "ExcelUpdater":
        self.wb = openpyxl.load_workbook(self.excel_path)
        return self

    def save(self):
        if self.wb:
            self.wb.save(self.excel_path)

    def get_sheet(self, sheet_name: str):
        if not self.wb:
            raise ExcelWriteError("Workbook no cargado")
        return self.wb[sheet_name]

    def find_and_write_entry(
        self,
        vehicle_name: str,
        entry: DailyEntry
    ) -> tuple[bool, str]:
        if not self.wb:
            return False, "Workbook no cargado"

        from .finder import SheetFinder, find_column_by_header

        sheet_finder = SheetFinder(self.wb)
        sheet = sheet_finder.find_vehicle_sheet(entry.vehicle.name, entry.vehicle.plate)

        if not sheet:
            return False, f"Hoja no encontrada para {entry.vehicle.name}"

        headers = find_column_by_header(sheet, ["FECHA", "KILOMETRAJE", "EXCESO", "ESTACIONAMIENTO", "COMBUSTIBLE"])

        writer = SafeExcelWriter(self.excel_path)
        writer.wb = self.wb

        try:
            success = writer.write_entry_safe(sheet, entry, headers)
            if success:
                self.save()
            return success, "OK" if success else "Duplicado"
        except Exception as e:
            return False, str(e)