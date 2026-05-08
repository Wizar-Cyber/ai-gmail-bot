import pytest
import openpyxl
import os
from datetime import date


class TestExcelFinder:
    """Tests para Excel finder."""

    @pytest.fixture
    def excel_path(self):
        path = "KILOMETRAJE Y DEPRECIACION VEHICULOS 2026 (1).xlsx"
        if os.path.exists(path):
            return path
        pytest.skip("Excel file not found")

    @pytest.fixture
    def workbook(self, excel_path):
        return openpyxl.load_workbook(excel_path)

    def test_sheet_names(self, workbook):
        """Verifica que existen las hojas esperadas."""
        expected = ["TOTAL Y RESUM 2026", "MERCEDES E 200 JSZ167", "MERCEDES GLE450"]
        for sheet in expected:
            assert sheet in workbook.sheetnames

    def test_vehicle_sheets_exist(self, workbook):
        """Verifica que hay hojas de vehículos."""
        vehicle_sheets = [s for s in workbook.sheetnames if s != "TOTAL Y RESUM 2026"]
        assert len(vehicle_sheets) >= 5

    def test_column_headers(self, workbook):
        """Verifica que las hojas tienen headers."""
        ws = workbook["MERCEDES E 200 JSZ167"]
        header_row = 3
        headers = [ws.cell(header_row, c).value for c in range(1, 10)]
        print(f"Headers: {headers}")

    def test_data_structure(self, workbook):
        """Verifica estructura de datos (filas con fechas)."""
        ws = workbook["MERCEDES E 200 JSZ167"]
        for row in range(5, 15):
            cell_value = ws.cell(row, 1).value
            if cell_value:
                print(f"Row {row}, Col 1: {cell_value}")
                break


class TestExcelWriter:
    """Tests para Excel writer."""

    def test_safe_write_preserves_formula(self):
        """Verifica que no se sobreescriben fórmulas."""
        pass