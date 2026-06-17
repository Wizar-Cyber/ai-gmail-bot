import pytest
import os


class TestIntegration:
    """Tests de integración del pipeline completo."""

    def test_pipeline_with_existing_pdf(self):
        """Testea el pipeline completo con el PDF existente."""
        from src.pdf.extractor import extract_reports_from_pdf
        from src.excel.writer import ExcelUpdater
        from src.storage.database import Database

        pdf_path = "Rapport kilométrique(20260423-20260423).pdf"
        excel_path = "KILOMETRAJE Y DEPRECIACION VEHICULOS 2026 (1).xlsx"

        if not os.path.exists(pdf_path):
            pytest.skip("PDF not found")
        if not os.path.exists(excel_path):
            pytest.skip("Excel not found")

        reports = extract_reports_from_pdf(pdf_path)
        print(f"Extracted {len(reports)} reports from PDF")

        assert len(reports) > 0

        for report in reports:
            print(f"Vehicle: {report.vehicle.name}")
            print(f"  Entries: {len(report.entries)}")
            for entry in report.entries:
                print(f"    {entry.date}: {entry.kilometers}km, {entry.fuel_liters}L")

    def test_database_operations(self):
        """Testea operaciones de base de datos."""
        from src.storage.database import Database
        import os
        test_db = "sqlite:///data/test_integration.db"
        if os.path.exists("data/test_integration.db"):
            os.remove("data/test_integration.db")
        db = Database(test_db)

        result = db.save_entry(
            vehicle_name="MERCEDES E200",
            vehicle_plate="JSZ167",
            entry_date="2026-04-23",
            kilometers=114.962,
            speed_excess=0,
            parking_time=28,
            fuel=9.2,
            source_file="test.pdf",
            email_id="test123"
        )
        assert result is True

        entries = db.get_entries_by_date("2026-04-01", "2026-04-30")
        print(f"Entries in date range: {len(entries)}")
        assert len(entries) >= 1

    def test_excel_finder_with_real_file(self):
        """Testea finder con el Excel real."""
        excel_path = "KILOMETRAJE Y DEPRECIACION VEHICULOS 2026 (1).xlsx"
        if not os.path.exists(excel_path):
            pytest.skip("Excel not found")

        from src.excel.finder import SheetFinder
        import openpyxl

        wb = openpyxl.load_workbook(excel_path)
        finder = SheetFinder(wb)

        sheet = finder.find_vehicle_sheet("MERCEDES E200", "JSZ167")
        assert sheet is not None
        print(f"Found sheet: {sheet.title}")

    def test_vehicle_mapping_fuzzy_match(self):
        """Testea fuzzy matching de vehículos."""
        from src.normalization.vehicles import find_excel_sheet_name
        import openpyxl

        excel_path = "KILOMETRAJE Y DEPRECIACION VEHICULOS 2026 (1).xlsx"
        if not os.path.exists(excel_path):
            pytest.skip("Excel not found")

        wb = openpyxl.load_workbook(excel_path)
        sheets = wb.sheetnames

        result = find_excel_sheet_name("MERCEDES E200", sheets)
        print(f"Matched sheet: {result}")

        result2 = find_excel_sheet_name("MERCEDES GLE450", sheets)
        print(f"Matched sheet 2: {result2}")